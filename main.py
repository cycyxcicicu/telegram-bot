from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackContext
from telegram.ext import filters
from openpyxl import Workbook, load_workbook
import os
import asyncio
import requests
from bs4 import BeautifulSoup
import random

# Đặt token của bạn ở đây
YOUR_BOT_TOKEN = os.getenv("YOUR_BOT_TOKEN")  # Thay thế bằng token thực tế của bạn

# Thiết lập proxy
# Thiết lập proxy
proxies_list = [
    {
        'http': 'http://d3530cadb9:M91VxFDm@168.91.33.188:4444',
        'https': 'http://d3530cadb9:M91VxFDm@168.91.33.188:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@208.52.181.96:4444',
        'https': 'http://d3530cadb9:M91VxFDm@208.52.181.96:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@69.58.65.116:4444',
        'https': 'http://d3530cadb9:M91VxFDm@69.58.65.116:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@136.0.116.214:4444',
        'https': 'http://d3530cadb9:M91VxFDm@136.0.116.214:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@168.91.39.173:4444',
        'https': 'http://d3530cadb9:M91VxFDm@168.91.39.173:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@168.91.36.98:4444',
        'https': 'http://d3530cadb9:M91VxFDm@168.91.36.98:4444',
    },
    {
        'http': 'http://d3530cadb9:M91VxFDm@168.91.47.160:4444',
        'https': 'http://d3530cadb9:M91VxFDm@168.91.47.160:4444',
    }
]



# Tạo file Excel nếu không tồn tại
excel_file = "messages.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    wb.save(excel_file)

user_status = {}
user_messages = {}
user_exported_index = {}
user_timers = {}
user_titles = {}
user_file_status = {}  # Trạng thái file đã gửi

async def export_user_messages(update, context, user_id):
    user_file = f"{user_id}_messages.xlsx"
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = str(user_id)
    new_ws.append(["User ID", "Message", "Title",
                    "Image 1", "Image 2", "Image 3",
                    "Image 4", "Image 5", "Image 6",
                    "Image 7", "Image 8", "Image 9"])  # Thêm tiêu đề cho 9 cột hình ảnh

    if user_id in user_messages:
        for idx, msg in enumerate(user_messages[user_id][user_exported_index[user_id]:]):
            title = clean_string(user_titles[user_id][idx][0] if idx < len(user_titles[user_id]) else "Không có tiêu đề")
            img_urls = user_titles[user_id][idx][1] if idx < len(user_titles[user_id]) else []

            # Ghi vào file Excel, lưu tất cả các URL hình ảnh
            img_row = img_urls + [""] * (12 - len(img_urls))  # Thêm ô trống nếu có ít hơn 9 ảnh
            new_ws.append([user_id, msg, title] + img_row)  # Ghi vào Excel

    new_wb.save(user_file)

    with open(user_file, 'rb') as file:
        await context.bot.send_document(chat_id=update.message.chat.id, document=file)

async def stop_bot(update: Update, context: CallbackContext, user_id):
    await asyncio.sleep(1800)  # Thời gian chờ 30 phút
    if user_id in user_status and user_status[user_id]:
        if user_exported_index[user_id] < len(user_messages[user_id]):
            await export_user_messages(update, context, user_id)
        await update.message.reply_text("Bot sẽ dừng do không hoạt động trong 30 phút.")
        user_status[user_id] = False
        user_messages[user_id] = []
        user_titles[user_id] = []
        user_file_status[user_id] = False  # Đặt lại trạng thái file

async def start(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    user_status[user_id] = True
    user_messages[user_id] = []
    user_titles[user_id] = []  # Khởi tạo danh sách tiêu đề và ảnh
    user_exported_index[user_id] = 0
    user_file_status[user_id] = False  # Khởi tạo trạng thái file
    await update.message.reply_text('Xin chào! Tôi sẽ lưu tin nhắn của bạn vào file Excel từ bây giờ. Bắt đầu gửi tin nhắn!')

    if user_id in user_timers:
        user_timers[user_id].cancel()
    user_timers[user_id] = asyncio.create_task(stop_bot(update, context, user_id))

async def thongtin(update: Update, context: CallbackContext) -> None:
    commands_info = (
        "/start - Bắt đầu lưu tin nhắn.",
        "/export - Xuất tin nhắn đã lưu.",
        "/thongtin - Hiển thị thông tin về các lệnh hỗ trợ.",
        "/readfile - Đọc file Excel và lấy dữ liệu sản phẩm."
    )
    await update.message.reply_text("\n".join(commands_info))

async def export(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    await export_user_messages(update, context, user_id)
    user_exported_index[user_id] = len(user_messages[user_id])

async def fetch_url_data(url):
    proxy = random.choice(proxies_list)
    try:
        response = requests.get(url, proxies=proxy)
        response.raise_for_status()  # Kiểm tra mã trạng thái

        soup = BeautifulSoup(response.content, 'html.parser')

        # Lấy tiêu đề từ thẻ <div> có class 'index-title--AnTxK'
        title_div = soup.find('div', class_='index-title--AnTxK')
        title = title_div.get_text() if title_div else "Không tìm thấy tiêu đề."

        # Kiểm tra sự tồn tại của slick-track và lấy hình ảnh
        slick_track = soup.find('div', class_='slick-track')
        img_tags = slick_track.find_all('img') if slick_track else []
        img_urls = [img.get('src') for img in img_tags if img.get('src')]

        return title, img_urls

    except requests.exceptions.RequestException as e:
        print(f"Không thể truy cập trang: {e}")
        return "Không lấy được dữ liệu", []

async def echo(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    message_text = update.message.text

    if user_status.get(user_id, False):
        user_messages[user_id].append(message_text)

        # Kiểm tra nếu tin nhắn là một URL
        if message_text.startswith('http://') or message_text.startswith('https://'):
            title, img_urls = await fetch_url_data(message_text)
            user_titles[user_id].append((title, img_urls))
            num_images = len(img_urls)  # Đếm số lượng hình ảnh
            
            # Gửi thông báo lưu tin nhắn
            await update.message.reply_text(f'Tin nhắn của bạn đã được lưu: "{message_text}"\nTiêu đề: {title}\nSố lượng hình ảnh: {num_images}')
        else:
            user_titles[user_id].append(("Không phải link", []))
            await update.message.reply_text(f'Tin nhắn của bạn đã được lưu: "{message_text}"')

        # Thiết lập lại thời gian chờ
        if user_id in user_timers:
            user_timers[user_id].cancel()
        user_timers[user_id] = asyncio.create_task(stop_bot(update, context, user_id))
    else:
        await update.message.reply_text("Vui lòng sử dụng lệnh /start để bắt đầu lưu tin nhắn.")

async def read_excel_file(file_path):
    """Đọc file Excel và lấy các link sản phẩm từ cột 'Link'."""
    df = load_workbook(file_path)
    results = []

    for sheet in df.sheetnames:
        worksheet = df[sheet]
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua hàng tiêu đề
            url = row[0]  # Giả sử link nằm ở cột đầu tiên
            title, img_urls = await fetch_url_data(url)
            results.append({'URL': url, 'Title': title, 'Images': img_urls})

    return results
def clean_string(value):
    if isinstance(value, str):
        # Loại bỏ ký tự không hợp lệ
        return ''.join(char for char in value if char.isprintable())
    return value

async def read_file(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id

    # Kiểm tra trạng thái người dùng đã bắt đầu tương tác
    if not user_status.get(user_id, False):
        await update.message.reply_text("Vui lòng sử dụng lệnh /start để bắt đầu.")
        return

    # Kiểm tra nếu có file được gửi
    if update.message.document:
        file = await update.message.document.get_file()
        input_file = f"{user_id}_input_file.xlsx"
        await file.download_to_drive(input_file)  # Tải file về

        results = await read_excel_file(input_file)

        # Tạo file Excel mới với kết quả
        output_file = f"{user_id}_output.xlsx"
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.append(["URL", "Title"] + [f"Image URL {i+1}" for i in range(max(len(result['Images']) for result in results))])  # Tiêu đề cột

        for result in results:
            row = [result['URL'], clean_string(result['Title'])]
            row.extend(result['Images'])  # Thêm từng link hình ảnh vào hàng
            output_ws.append(row)

        output_wb.save(output_file)

        with open(output_file, 'rb') as f:
            await context.bot.send_document(chat_id=update.message.chat.id, document=f)

        # Đánh dấu là đã gửi file
        user_file_status[user_id] = True  # Đánh dấu là đã gửi file
        
        await update.message.reply_text("File đã được xử lý. Bạn có thể gửi file mới bất kỳ lúc nào.")
    else:
        await update.message.reply_text("Vui lòng gửi file Excel để xử lý.")

async def export(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    await export_user_messages(update, context, user_id)
    user_exported_index[user_id] = len(user_messages[user_id])
def main() -> None:
    application = ApplicationBuilder().token(YOUR_BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("thongtin", thongtin))
    application.add_handler(MessageHandler(filters.Document.ALL & ~filters.COMMAND, read_file))  # Xử lý file tải lên
    application.add_handler(CommandHandler("export", export))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, echo))

    application.run_polling()

if __name__ == '__main__':
    main()